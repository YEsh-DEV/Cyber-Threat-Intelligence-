"""
Professor Demonstration Runner v2
=================================
Uses Gemini for extraction (generous rate limits) and local Ollama qwen for evaluation.
Directly controls event loading to guarantee exactly 10 demo events are processed.
"""

import json
import os
import sys
import time
import random
import shutil
import csv
import logging
from pathlib import Path
from datetime import datetime

# Setup paths
PROJECT_ROOT = Path("Y:/Reserchintern/Experiment2")
sys.path.insert(0, str(PROJECT_ROOT))

# Setup logging
(PROJECT_ROOT / "logs").mkdir(parents=True, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(PROJECT_ROOT / "logs" / "demo_run.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("demo_runner")

# Configuration
EXTRACTION_MODEL = "llama_groq"      # Groq API for extraction
EVALUATOR_MODEL = "ollama_gemma"      # Local Ollama - no API quota
METHODS = ["llm_only", "vanilla_rag", "graph_rag"]
NUM_EVENTS = 10
RANDOM_SEED = 42
DEMO_DIR = PROJECT_ROOT / "outputs" / "demo_run"


def get_demo_events():
    """Load or create the 10 demo events."""
    demo_path = DEMO_DIR / "demo_events.json"
    if demo_path.exists():
        with open(demo_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def step2_create_demo_dataset():
    """Step 2: Create demo_events.json with 10 semantic events."""
    print("\n" + "=" * 70)
    print("  STEP 2: Creating Demo Dataset (10 Semantic Events)")
    print("=" * 70)

    from preprocessing.preprocess import load_cached_events

    all_events = load_cached_events()
    logger.info("Total events in cache: %d", len(all_events))

    # Filter to semantic events
    semantic_events = [
        e for e in all_events
        if len(e.get("narrative", "")) > 150 and len(e.get("narrative", "").split()) > 20
    ]
    logger.info("Semantic events available: %d", len(semantic_events))

    # Deterministic sampling
    semantic_events = sorted(semantic_events, key=lambda x: x.get("global_id", ""))
    random.seed(RANDOM_SEED)
    demo_events = random.sample(semantic_events, min(NUM_EVENTS, len(semantic_events)))

    # Save demo dataset
    DEMO_DIR.mkdir(parents=True, exist_ok=True)
    with open(DEMO_DIR / "demo_events.json", "w", encoding="utf-8") as f:
        json.dump(demo_events, f, indent=2, ensure_ascii=False)

    # Generate demo_dataset_report.md
    report_lines = [
        "# Demo Dataset Report\n",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Total Events Selected:** {len(demo_events)}",
        f"**Random Seed:** {RANDOM_SEED}\n",
        "## Selected Events\n",
        "| # | Global ID | Source File | Narrative Length (chars) | Word Count |",
        "|---|-----------|-------------|------------------------|------------|",
    ]
    for i, evt in enumerate(demo_events, 1):
        narr = evt.get("narrative", "")
        report_lines.append(
            f"| {i} | {evt['global_id']} | {evt['file_source']} | {len(narr)} | {len(narr.split())} |"
        )
    report_lines.append(f"\n## Sample Narrative (Event 1)\n")
    report_lines.append(f"```\n{demo_events[0]['narrative'][:500]}...\n```\n")

    with open(DEMO_DIR / "demo_dataset_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))

    print(f"  [OK] Created demo_events.json with {len(demo_events)} events")
    print(f"  [OK] Generated demo_dataset_report.md")
    return demo_events


def step3_run_single_method(method, demo_events):
    """Run extraction for a single method on demo events."""
    from models.llm_factory import LLMFactory
    from retrievers.retriever_factory import RetrieverFactory
    from schemas.extraction_schema import ExtractionResult

    print(f"\n  --- Running: {EXTRACTION_MODEL} + {method} ---")

    method_dir = DEMO_DIR / method
    method_dir.mkdir(parents=True, exist_ok=True)

    # Initialize components
    model = LLMFactory.create(EXTRACTION_MODEL)
    retriever = RetrieverFactory.create(method)

    # Load prompt template
    from config import EXTRACTION_PROMPT_FILE
    prompt_template = EXTRACTION_PROMPT_FILE.read_text(encoding="utf-8")
    system_prompt = prompt_template.replace("{context_block}", "").replace("{event_narrative}", "").strip()

    results = []
    start_time = time.time()

    for i, event in enumerate(demo_events):
        event_start = time.time()
        global_id = event["global_id"]
        narrative = event["narrative"]

        logger.info("Processing event %d/%d (Global ID=%s) [%s]", i + 1, len(demo_events), global_id, method)

        try:
            # Step 1: Get retrieval context
            retriever_start = time.time()
            is_ioc_only = len(narrative) < 150 or len(narrative.split()) < 20

            if method != "llm_only" and not is_ioc_only:
                context = retriever.get_context(narrative, global_id=global_id)
            else:
                context = []
            retriever_latency = time.time() - retriever_start

            # Step 2: Build prompt
            if context:
                context_text = "====================\nBACKGROUND CONTEXT\n(Not Ground Truth)\n====================\n"
                for ci, passage in enumerate(context, 1):
                    context_text += f"\n### Context {ci}\n{passage}\n"
            else:
                context_text = ""
            user_prompt = f"{context_text}\n====================\nEVENT NARRATIVE\n(Only Source of Truth)\n====================\n{narrative}"

            # Step 3: Query LLM
            model_start = time.time()
            raw_result = model.generate_json(system_prompt, user_prompt)
            model_latency = time.time() - model_start

            # Step 4: Basic validation
            entities = raw_result.get("entities", [])
            relations = raw_result.get("relations", [])

            processing_time = time.time() - event_start

            results.append({
                "global_id": global_id,
                "event_id": event["event_id"],
                "file_source": event["file_source"],
                "extraction": raw_result,
                "processing_time_seconds": round(processing_time, 2),
                "model_latency_seconds": round(model_latency, 2),
                "retriever_latency_seconds": round(retriever_latency, 2),
                "status": "success",
            })
            print(f"    [{i+1}/{len(demo_events)}] {global_id}: {len(entities)} entities, {len(relations)} relations ({processing_time:.1f}s)")

            # Rate limit protection - Groq free tier is 6000 TPM
            # Each request ~2000-3000 tokens, so wait 30s between calls
            time.sleep(30)

        except Exception as e:
            processing_time = time.time() - event_start
            logger.error("Error processing event %s: %s", global_id, e)
            results.append({
                "global_id": global_id,
                "event_id": event["event_id"],
                "file_source": event["file_source"],
                "extraction": {},
                "processing_time_seconds": round(processing_time, 2),
                "status": "error",
                "error_message": str(e),
            })
            print(f"    [{i+1}/{len(demo_events)}] {global_id}: ERROR - {e}")
            time.sleep(15)

    total_time = time.time() - start_time

    # Save output
    output = {
        "experiment_metadata": {
            "method": method,
            "model": EXTRACTION_MODEL,
            "timestamp": datetime.now().isoformat(),
            "dataset_size": len(results),
            "dev_mode": False,
            "run_id": f"demo_{method}_{EXTRACTION_MODEL}",
        },
        "results": results,
    }

    output_path = DEMO_DIR / f"extraction_{method}.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    success_count = sum(1 for r in results if r["status"] == "success")
    print(f"  [OK] {method}: {success_count}/{len(results)} succeeded in {total_time:.1f}s")

    return {
        "output_path": str(output_path),
        "runtime": total_time,
        "status": "success",
        "success_count": success_count,
        "total_count": len(results),
    }


def step3_run_methods(demo_events):
    """Step 3: Run all three methods."""
    print("\n" + "=" * 70)
    print("  STEP 3: Running Three Extraction Methods")
    print("=" * 70)

    results = {}
    for method in METHODS:
        results[method] = step3_run_single_method(method, demo_events)

    return results


def step4_evaluate(extraction_results):
    """Step 4: Run evaluation using local Ollama qwen model."""
    print("\n" + "=" * 70)
    print(f"  STEP 4: Running Evaluation (Judge: {EVALUATOR_MODEL})")
    print("=" * 70)

    from evaluation.evaluator import Evaluator
    evaluator = Evaluator(evaluator_model_name=EVALUATOR_MODEL)
    eval_results = {}

    for method in METHODS:
        info = extraction_results.get(method, {})
        if info.get("status") != "success":
            print(f"  [SKIP] Skipping {method} (extraction failed)")
            eval_results[method] = None
            continue

        extraction_file = DEMO_DIR / f"extraction_{method}.json"
        if not extraction_file.exists():
            print(f"  [SKIP] Skipping {method}: extraction file not found")
            eval_results[method] = None
            continue

        print(f"\n  --- Evaluating: {method} ---")
        try:
            report = evaluator.evaluate_batch(str(extraction_file))
            eval_results[method] = report

            eval_path = DEMO_DIR / f"evaluation_{method}.json"
            with open(eval_path, "w", encoding="utf-8") as f:
                json.dump(report, f, indent=2, ensure_ascii=False)

            stats = report.get("statistics", {})
            avgs = stats.get("averages", {})
            print(f"  [OK] {method} evaluated: faith={avgs.get('faithfulness', 'N/A')}, "
                  f"rel={avgs.get('relevance', 'N/A')}, "
                  f"halluc={avgs.get('hallucination_rate', 'N/A')}")

        except Exception as e:
            logger.error("Evaluation of %s failed: %s", method, e)
            import traceback
            traceback.print_exc()
            eval_results[method] = None
            print(f"  [FAIL] {method} evaluation FAILED: {e}")

    return eval_results


def step5_comparison_table(extraction_results, eval_results):
    """Step 5: Generate comparison table."""
    print("\n" + "=" * 70)
    print("  STEP 5: Generating Comparison Table")
    print("=" * 70)

    rows = []
    for method in METHODS:
        ext_info = extraction_results.get(method, {})
        ev_info = eval_results.get(method)

        entity_count = 0
        relation_count = 0
        success_count = 0
        total_count = 0

        extraction_file = DEMO_DIR / f"extraction_{method}.json"
        if extraction_file.exists():
            with open(extraction_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            for r in data.get("results", []):
                total_count += 1
                if r.get("status") in ("success", "partial"):
                    success_count += 1
                    extraction = r.get("extraction", {})
                    entity_count += len(extraction.get("entities", []))
                    relation_count += len(extraction.get("relations", []))

        faithfulness = "N/A"
        relevance = "N/A"
        coverage = "N/A"
        hallucination = "N/A"

        if ev_info:
            avgs = ev_info.get("statistics", {}).get("averages", {})
            if avgs.get("faithfulness") is not None:
                faithfulness = f"{avgs['faithfulness']:.3f}"
            if avgs.get("relevance") is not None:
                relevance = f"{avgs['relevance']:.3f}"
            if avgs.get("evidence_coverage") is not None:
                coverage = f"{avgs['evidence_coverage']:.3f}"
            if avgs.get("hallucination_rate") is not None:
                hallucination = f"{avgs['hallucination_rate']:.3f}"

        success_rate = f"{(success_count / total_count * 100):.1f}%" if total_count > 0 else "N/A"
        runtime = f"{ext_info.get('runtime', 0):.1f}s" if ext_info.get("status") == "success" else "N/A"

        rows.append({
            "Method": method,
            "Model": EXTRACTION_MODEL,
            "Runtime": runtime,
            "Success Rate": success_rate,
            "Entity Count": entity_count,
            "Relation Count": relation_count,
            "Faithfulness": faithfulness,
            "Relevance": relevance,
            "Coverage": coverage,
            "Hallucination": hallucination,
        })

    # CSV
    csv_path = DEMO_DIR / "benchmark_demo_table.csv"
    fieldnames = list(rows[0].keys()) if rows else []
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    # XLSX
    xlsx_path = DEMO_DIR / "benchmark_demo_table.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark Results"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, key in enumerate(fieldnames, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        for col in range(1, len(fieldnames) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(rows) + 2))
            col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) // 26) + chr(65 + (col - 1) % 26)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(xlsx_path)
        print(f"  [OK] Generated benchmark_demo_table.xlsx")
    except Exception as e:
        print(f"  [WARN] XLSX generation failed: {e}")

    print(f"  [OK] Generated benchmark_demo_table.csv")

    # Summary report
    summary_lines = [
        "# Demo Benchmark Summary Report\n",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Extraction Model:** {EXTRACTION_MODEL}",
        f"**Evaluator Model:** {EVALUATOR_MODEL}",
        f"**Dataset:** {NUM_EVENTS} semantic CTI events\n",
        "## Results\n",
        "| Method | Model | Runtime | Success Rate | Entities | Relations | Faithfulness | Relevance | Coverage | Hallucination |",
        "|--------|-------|---------|-------------|----------|-----------|-------------|-----------|----------|---------------|",
    ]
    for row in rows:
        summary_lines.append(
            f"| {row['Method']} | {row['Model']} | {row['Runtime']} | {row['Success Rate']} | "
            f"{row['Entity Count']} | {row['Relation Count']} | {row['Faithfulness']} | "
            f"{row['Relevance']} | {row['Coverage']} | {row['Hallucination']} |"
        )

    with open(DEMO_DIR / "summary_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(summary_lines))

    print(f"  [OK] Generated summary_report.md")
    return rows


def step6_best_method(rows):
    """Step 6: Select best performing method."""
    print("\n" + "=" * 70)
    print("  STEP 6: Selecting Best Performing Method")
    print("=" * 70)

    best = None
    best_score = -1
    for row in rows:
        score = row.get("Entity Count", 0) + row.get("Relation Count", 0)
        if score > best_score:
            best_score = score
            best = row

    if not best:
        print("  [FAIL] No successful method found!")
        return None

    best_method = best["Method"]
    print(f"  [OK] Best method: {best_method} (score: {best_score} entities+relations)")

    # Sample extraction
    sample_extraction = {}
    extraction_file = DEMO_DIR / f"extraction_{best_method}.json"
    if extraction_file.exists():
        with open(extraction_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for r in data.get("results", []):
            if r.get("status") in ("success", "partial") and r.get("extraction", {}).get("entities"):
                sample_extraction = r
                break

    report_lines = [
        "# Best Method Report\n",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Best Method:** `{best_method}`",
        f"**Extraction Model:** `{EXTRACTION_MODEL}`",
        f"**Evaluator Model:** `{EVALUATOR_MODEL}`\n",
        "## Performance Summary\n",
        f"- **Runtime:** {best['Runtime']}",
        f"- **Success Rate:** {best['Success Rate']}",
        f"- **Entity Count:** {best['Entity Count']}",
        f"- **Relation Count:** {best['Relation Count']}\n",
        "## Evaluation Metrics\n",
        f"- **Faithfulness:** {best['Faithfulness']}",
        f"- **Relevance:** {best['Relevance']}",
        f"- **Evidence Coverage:** {best['Coverage']}",
        f"- **Hallucination Rate:** {best['Hallucination']}\n",
    ]

    if sample_extraction:
        sample_json = json.dumps(sample_extraction.get("extraction", {}), indent=2, ensure_ascii=False)
        report_lines.extend([
            "## Sample Extraction Output\n",
            f"**Event:** {sample_extraction.get('global_id', 'N/A')}\n",
            f"```json\n{sample_json[:3000]}\n```\n",
        ])

    with open(DEMO_DIR / "best_method_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))

    print(f"  [OK] Generated best_method_report.md")
    return best_method


def step7_neo4j_preparation(best_method):
    """Step 7: Neo4j preparation."""
    print("\n" + "=" * 70)
    print("  STEP 7: Neo4j Preparation")
    print("=" * 70)

    extraction_file = DEMO_DIR / f"extraction_{best_method}.json"
    entity_count = 0
    relation_count = 0

    if extraction_file.exists():
        with open(extraction_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for r in data.get("results", []):
            if r.get("status") in ("success", "partial"):
                ext = r.get("extraction", {})
                entity_count += len(ext.get("entities", []))
                relation_count += len(ext.get("relations", []))

    report_lines = [
        "# Neo4j Ready Report\n",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Best Method:** `{best_method}`",
        f"**Extraction Model:** `{EXTRACTION_MODEL}`\n",
        "## Data Summary\n",
        f"- **JSON Path:** `{extraction_file}`",
        f"- **Entity Count:** {entity_count}",
        f"- **Relation Count:** {relation_count}",
        f"- **Estimated Nodes (Event + Entity):** ~{10 + entity_count}",
        f"- **Estimated Edges (HAS_ENTITY + Relations):** ~{entity_count + relation_count}\n",
        "## Neo4j Connection\n",
        "- **URI:** `neo4j+ssc://d3364235.databases.neo4j.io`",
        "- **Database:** `d3364235`\n",
    ]

    try:
        from graph.neo4j_loader import Neo4jLoader
        loader = Neo4jLoader()
        loader.connect()
        stats = loader.get_graph_stats()
        loader.close()
        report_lines.extend([
            "## Connection Status: SUCCESS\n",
            f"- **Existing Nodes:** {stats.get('node_counts', {})}",
            f"- **Existing Relationships:** {stats.get('total_relationships', 0)}\n",
        ])
        print(f"  [OK] Neo4j connection verified")
    except Exception as e:
        report_lines.append(f"## Connection Status: FAILED ({e})\n")
        print(f"  [FAIL] Neo4j connection failed: {e}")

    with open(DEMO_DIR / "neo4j_ready_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))
    print(f"  [OK] Generated neo4j_ready_report.md")


def step8_neo4j_import(best_method):
    """Step 8: Import into Neo4j."""
    print("\n" + "=" * 70)
    print("  STEP 8: Neo4j Import")
    print("=" * 70)

    extraction_file = DEMO_DIR / f"extraction_{best_method}.json"
    if not extraction_file.exists():
        print(f"  [FAIL] Extraction file not found: {extraction_file}")
        return

    report_lines = [
        "# Neo4j Import Report\n",
        f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**Source:** `{extraction_file}`",
        f"**Method:** `{best_method}`\n",
    ]

    try:
        from graph.neo4j_loader import Neo4jLoader

        start_time = time.time()
        with Neo4jLoader() as loader:
            stats = loader.load_json(str(extraction_file), append=True)
            runtime = time.time() - start_time

            report_lines.extend([
                "## Import Results: SUCCESS\n",
                f"- **Events Created:** {stats['events_created']}",
                f"- **Entities Created:** {stats['entities_created']}",
                f"- **Relations Created:** {stats['relations_created']}",
                f"- **Import Runtime:** {runtime:.2f}s\n",
            ])

            # Verification queries
            from neo4j import GraphDatabase
            from config import NEO4J_URI, NEO4J_USERNAME, NEO4J_PASSWORD, NEO4J_DATABASE

            driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))
            with driver.session(database=NEO4J_DATABASE) as session:
                r1 = session.run("MATCH (n) RETURN count(n) AS count").single()["count"]
                r2 = session.run("MATCH ()-[r]->() RETURN count(r) AS count").single()["count"]

                r3 = session.run("MATCH (n) RETURN labels(n) AS labels, count(n) AS cnt ORDER BY cnt DESC")
                label_rows = [(str(r["labels"]), r["cnt"]) for r in r3]

                r4 = session.run("MATCH ()-[r]->() RETURN type(r) AS type, count(r) AS cnt ORDER BY cnt DESC LIMIT 10")
                rel_type_rows = [(r["type"], r["cnt"]) for r in r4]

            driver.close()

            report_lines.extend([
                "## Verification Queries\n",
                "```cypher",
                f"MATCH (n) RETURN count(n);  -- Result: {r1}",
                f"MATCH ()-[r]->() RETURN count(r);  -- Result: {r2}",
                "```\n",
                "### Node Labels\n",
                "| Label | Count |",
                "|-------|-------|",
            ])
            for label, cnt in label_rows:
                report_lines.append(f"| {label} | {cnt} |")

            report_lines.extend(["\n### Relationship Types\n", "| Type | Count |", "|------|-------|"])
            for rtype, cnt in rel_type_rows:
                report_lines.append(f"| {rtype} | {cnt} |")

        print(f"  [OK] Imported: {stats['events_created']} events, {stats['entities_created']} entities, {stats['relations_created']} relations")
        print(f"  [OK] Graph: {r1} nodes, {r2} relationships")

    except Exception as e:
        logger.error("Neo4j import failed: %s", e)
        import traceback
        traceback.print_exc()
        report_lines.append(f"## Import Results: FAILED ({e})\n")
        print(f"  [FAIL] Neo4j import failed: {e}")

    with open(DEMO_DIR / "neo4j_import_report.md", "w", encoding="utf-8") as f:
        f.write("\n".join(report_lines))
    print(f"  [OK] Generated neo4j_import_report.md")


def step9_demo_package(best_method):
    """Step 9: Assemble demo package."""
    print("\n" + "=" * 70)
    print("  STEP 9: Assembling Demo Package")
    print("=" * 70)

    pkg_dir = DEMO_DIR / "demo_package"
    pkg_dir.mkdir(parents=True, exist_ok=True)

    files_to_copy = [
        ("benchmark_demo_table.csv", "benchmark_demo_table.csv"),
        ("benchmark_demo_table.xlsx", "benchmark_demo_table.xlsx"),
        ("best_method_report.md", "best_method_report.md"),
        ("neo4j_import_report.md", "neo4j_import_report.md"),
        ("summary_report.md", "summary_report.md"),
        ("demo_dataset_report.md", "demo_dataset_report.md"),
        ("neo4j_ready_report.md", "neo4j_ready_report.md"),
        (f"extraction_{best_method}.json", "sample_extractions.json"),
    ]

    for src_name, dest_name in files_to_copy:
        src = DEMO_DIR / src_name
        if src.exists():
            shutil.copy2(src, pkg_dir / dest_name)
            print(f"  [OK] Copied {dest_name}")

    # screenshots_needed.md
    with open(pkg_dir / "screenshots_needed.md", "w", encoding="utf-8") as f:
        f.write("\n".join([
            "# Screenshots Needed for Presentation\n",
            "1. Terminal output showing all steps completing",
            "2. Neo4j Browser: `MATCH (n)-[r]->(m) RETURN n,r,m LIMIT 50`",
            "3. Excel: `benchmark_demo_table.xlsx`",
            "4. JSON: `sample_extractions.json`\n",
            "## Neo4j Console: https://console.neo4j.io\n",
            "## Useful Cypher Queries\n",
            "```cypher",
            "MATCH (n)-[r]->(m) RETURN n,r,m LIMIT 100",
            "MATCH (e:Event)-[:HAS_ENTITY]->(ent:Entity) RETURN e.event_id, ent.canonical_name, ent.type LIMIT 20",
            "MATCH (ent:Entity) RETURN ent.type, count(ent) AS count ORDER BY count DESC",
            "```\n",
        ]))

    print(f"  [OK] Demo package assembled at: {pkg_dir}")


def step10_walkthrough():
    """Step 10: Generate demo walkthrough."""
    print("\n" + "=" * 70)
    print("  STEP 10: Generating Demo Walkthrough")
    print("=" * 70)

    with open(DEMO_DIR / "demo_walkthrough.md", "w", encoding="utf-8") as f:
        f.write("\n".join([
            "# Professor Demonstration Walkthrough\n",
            "## Prerequisites\n",
            "- Python venv: `Y:\\Reserchintern\\Experiment1\\.venv`",
            "- Ollama running locally with `qwen2.5-coder:7b`",
            "- Internet for Gemini API",
            "- Neo4j AuraDB active\n",
            "## Quick Start\n",
            "```powershell",
            "Y:\\Reserchintern\\Experiment1\\.venv\\Scripts\\Activate.ps1",
            "cd Y:\\Reserchintern\\Experiment2",
            "$env:PYTHONUTF8='1'",
            "python run_demo.py",
            "```\n",
            "## Step-by-Step Commands\n",
            "### 1. Activate Environment\n",
            "```powershell",
            "Y:\\Reserchintern\\Experiment1\\.venv\\Scripts\\Activate.ps1",
            "cd Y:\\Reserchintern\\Experiment2",
            "```\n",
            "### 2. Preprocess Dataset\n",
            "```powershell",
            "python -m preprocessing.preprocess",
            "```\n",
            "### 3. Run Full Demo Pipeline\n",
            "```powershell",
            "$env:PYTHONUTF8='1'; python run_demo.py",
            "```\n",
            "### 4. View Results\n",
            "```powershell",
            "dir outputs/demo_run/",
            "```\n",
            "### 5. Open Neo4j Console\n",
            "Visit: https://console.neo4j.io\n",
            "```cypher",
            "MATCH (n)-[r]->(m) RETURN n,r,m LIMIT 50",
            "```\n",
            "## Output Files\n",
            "| File | Description |",
            "|------|-------------|",
            "| `demo_events.json` | 10 selected CTI events |",
            "| `extraction_llm_only.json` | LLM-only extraction |",
            "| `extraction_vanilla_rag.json` | Vanilla RAG extraction |",
            "| `extraction_graph_rag.json` | GraphRAG extraction |",
            "| `evaluation_*.json` | Evaluation scores |",
            "| `benchmark_demo_table.xlsx` | Comparison table |",
            "| `best_method_report.md` | Best method analysis |",
            "| `neo4j_import_report.md` | Neo4j import verification |",
        ]))

    print(f"  [OK] Generated demo_walkthrough.md")


def main():
    print("\n" + "#" * 70)
    print("  PROFESSOR DEMONSTRATION BUILD")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Extraction Model: {EXTRACTION_MODEL}")
    print(f"  Evaluator Model: {EVALUATOR_MODEL}")
    print(f"  Methods: {', '.join(METHODS)}")
    print(f"  Events: {NUM_EVENTS}")
    print("#" * 70)

    overall_start = time.time()

    # Step 2
    demo_events = step2_create_demo_dataset()

    # Step 3
    extraction_results = step3_run_methods(demo_events)

    # Step 4
    eval_results = step4_evaluate(extraction_results)

    # Step 5
    rows = step5_comparison_table(extraction_results, eval_results)

    # Step 6
    best_method = step6_best_method(rows)

    if best_method:
        step7_neo4j_preparation(best_method)
        step8_neo4j_import(best_method)
        step9_demo_package(best_method)

    step10_walkthrough()

    total_time = time.time() - overall_start
    print("\n" + "#" * 70)
    print(f"  DEMONSTRATION BUILD COMPLETE")
    print(f"  Total time: {total_time:.1f}s ({total_time/60:.1f} minutes)")
    print(f"  Output directory: {DEMO_DIR}")
    print(f"  Best method: {best_method or 'N/A'}")
    print("#" * 70)


if __name__ == "__main__":
    main()
